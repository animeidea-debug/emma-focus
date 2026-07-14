#!/usr/bin/env python3
"""
Emma Focus — XLSX → SQLite 恢复脚本
====================================
将 Google Sheet 导出的 Emma_Focus_DB.xlsx 导入本地 SQLite 数据库。
同时生成一份 JSON 备份供未来使用。

Usage:
    python3 deploy/import_xlsx_to_sqlite.py <xlsx_path> [db_path]

Default db_path: infra/web/backend/data/poc.db (本地开发)
For NAS: /app/data/poc.db
"""
import openpyxl
import sqlite3
import sys
import os
import json
from datetime import datetime

# ─── Category → Bucket mapping (mirrors poc_main.py) ───────────────────
CATEGORY_BUCKETS = {
    "Focus":       "focus",
    "Coaching":    "coaching",
    "Screen":      "screen",
    "Activity":    "activity",
    "Eye Rest":    "eyerest",
    "Distraction": "distraction",
}

def bucket_for(category):
    return CATEGORY_BUCKETS.get(str(category).strip(), "neutral")

def ensure_tables(conn):
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS evaluations (
            date TEXT PRIMARY KEY,
            day_type TEXT DEFAULT '',
            start_time TEXT DEFAULT '',
            end_time TEXT DEFAULT '',
            focus_blocks INTEGER DEFAULT 0,
            distractions INTEGER DEFAULT 0,
            eye_rest_minutes INTEGER DEFAULT 0,
            rating TEXT DEFAULT '',
            summary TEXT DEFAULT '',
            note TEXT DEFAULT '',
            absent INTEGER DEFAULT 0,
            status TEXT DEFAULT 'gray',
            tokens_net INTEGER DEFAULT 0,
            bucket_focus INTEGER DEFAULT 0,
            bucket_coaching INTEGER DEFAULT 0,
            bucket_screen INTEGER DEFAULT 0,
            bucket_activity INTEGER DEFAULT 0,
            bucket_eyerest INTEGER DEFAULT 0,
            bucket_distraction INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS activity_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            stage_name TEXT DEFAULT '',
            category TEXT DEFAULT '',
            duration INTEGER DEFAULT 0,
            start_time TEXT DEFAULT '',
            end_time TEXT DEFAULT '',
            note TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS tokens (
            silver_balance INTEGER DEFAULT 0,
            gold_balance INTEGER DEFAULT 0,
            exchange_rate INTEGER DEFAULT 5
        );
        CREATE TABLE IF NOT EXISTS token_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            type TEXT NOT NULL,
            description TEXT DEFAULT '',
            silver_delta INTEGER DEFAULT 0,
            gold_delta INTEGER DEFAULT 0,
            silver_balance INTEGER DEFAULT 0,
            gold_balance INTEGER DEFAULT 0,
            note TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS redeem_items (
            item_id TEXT PRIMARY KEY,
            label TEXT NOT NULL,
            description TEXT DEFAULT '',
            coin_type TEXT NOT NULL,
            cost INTEGER NOT NULL,
            active INTEGER DEFAULT 1,
            sort_order INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS app_config (
            key TEXT PRIMARY KEY,
            value TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS api_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT NOT NULL,
            action TEXT NOT NULL,
            ip TEXT DEFAULT '',
            payload TEXT DEFAULT ''
        );
        INSERT OR IGNORE INTO tokens (silver_balance, gold_balance, exchange_rate) VALUES (0, 0, 5);
        INSERT OR IGNORE INTO app_config (key, value) VALUES ('exchange_rate', '5');
    """)
    conn.commit()

def safe_int(v):
    try: return int(float(v)) if v is not None else 0
    except: return 0

def safe_str(v):
    if v is None: return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    try:
        import datetime as dt_mod
        if isinstance(v, dt_mod.time):
            return v.strftime("%H:%M")
    except:
        pass
    return str(v).strip()

def safe_bool(v):
    if v is None: return False
    v = str(v).strip().lower()
    return v in ("true", "1", "yes", "是")

def import_sheet(conn, wb):
    """Import all sheets from workbook"""
    sheet_map = {ws.title: ws for ws in wb.worksheets}
    
    # 1. Timeline → evaluations
    if "Timeline" in sheet_map:
        ws = sheet_map["Timeline"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        print(f"📋 Timeline headers: {headers}")
        print(f"📋 Timeline total rows (excl header): {len(rows)}")
        
        count = 0
        for row in rows:
            if not any(row): continue
            # Try positional (GAS export format)
            d = safe_str(row[0] if len(row) > 0 else "")
            if not d: continue  # skip empty dates
            
            day_type = safe_str(row[1] if len(row) > 1 else "")
            start_t  = safe_str(row[2] if len(row) > 2 else "")
            end_t    = safe_str(row[3] if len(row) > 3 else "")
            category = safe_str(row[4] if len(row) > 4 else "")
            fb       = safe_int(row[5] if len(row) > 5 else 0)
            dist     = safe_int(row[6] if len(row) > 6 else 0)
            note     = safe_str(row[7] if len(row) > 7 else "")
            absent   = safe_bool(row[8] if len(row) > 8 else False)
            eye      = safe_int(row[9] if len(row) > 9 else 0)
            
            absent_val = 1 if absent else 0
            status = "gray" if absent else "green"
            
            conn.execute("""
                INSERT OR REPLACE INTO evaluations
                (date, day_type, start_time, end_time, focus_blocks, distractions,
                 eye_rest_minutes, note, absent, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (d, day_type, start_t, end_t, fb, dist, eye, note, absent_val, status))
            count += 1
            
            # Also add to activity_logs if category is present
            if category:
                b = bucket_for(category)
                if b != "neutral":
                    conn.execute("""
                        INSERT INTO activity_logs (date, stage_name, category, duration, start_time, end_time, note)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (d, f"Timeline: {note[:20]}", category, 60 * fb if b == "focus" else 0, start_t, end_t, note))
        
        conn.commit()
        print(f"✅ Imported {count} timeline rows → evaluations")
    else:
        print("⚠️  No 'Timeline' sheet found")
    
    # 2. Evaluations → update rating, summary, tokens_net
    if "Evaluations" in sheet_map:
        ws = sheet_map["Evaluations"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        print(f"📋 Evaluations rows: {len(rows)}")
        
        count = 0
        for row in rows:
            if not any(row): continue
            d = safe_str(row[0] if len(row) > 0 else "")
            if not d: continue
            
            summary   = safe_str(row[1] if len(row) > 1 else "")
            rating    = safe_str(row[2] if len(row) > 2 else "")
            tokens_net = safe_int(row[3] if len(row) > 3 else 0)
            
            existing = conn.execute("SELECT * FROM evaluations WHERE date=?", (d,)).fetchone()
            if existing:
                conn.execute("""
                    UPDATE evaluations SET summary=?, rating=?, tokens_net=?
                    WHERE date=?
                """, (summary, rating, tokens_net, d))
            else:
                conn.execute("""
                    INSERT INTO evaluations (date, summary, rating, tokens_net, status)
                    VALUES (?, ?, ?, ?, 'green')
                """, (d, summary, rating, tokens_net))
            count += 1
            
            # Update status based on rating
            conn.execute("""
                UPDATE evaluations SET status = CASE
                    WHEN absent = 1 THEN 'gray'
                    WHEN rating LIKE '%🟢%' THEN 'green'
                    WHEN rating LIKE '%🟡%' THEN 'amber'
                    WHEN rating LIKE '%🔴%' THEN 'red'
                    WHEN rating LIKE '%⚪%' THEN 'gray'
                    ELSE status
                END WHERE date=?
            """, (d,))
        
        conn.commit()
        print(f"✅ Updated {count} evaluations")
    else:
        print("⚠️  No 'Evaluations' sheet found")
    
    # 3. Activity_Logs (XLSX columns: Date, Stage_Name, Start_Time, End_Time, Duration, Category, Note)
    if "Activity_Logs" in sheet_map:
        ws = sheet_map["Activity_Logs"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        print(f"📋 Activity_Logs rows: {len(rows)}")
        
        count = 0
        for row in rows:
            if not any(row): continue
            d        = safe_str(row[0] if len(row) > 0 else "")
            if not d: continue
            stage    = safe_str(row[1] if len(row) > 1 else "")
            start_t  = safe_str(row[2] if len(row) > 2 else "")
            end_t    = safe_str(row[3] if len(row) > 3 else "")
            dur      = safe_int(row[4] if len(row) > 4 else 0)
            category = safe_str(row[5] if len(row) > 5 else "")
            note     = safe_str(row[6] if len(row) > 6 else "")
            
            conn.execute("""
                INSERT INTO activity_logs (date, stage_name, category, duration, start_time, end_time, note)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (d, stage, category, dur, start_t, end_t, note))
            count += 1
        
        conn.commit()
        print(f"✅ Imported {count} activity logs")
    else:
        print("⚠️  No 'Activity_Logs' sheet found")
    
    # 4. Transactions
    if "Transactions" in sheet_map:
        ws = sheet_map["Transactions"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        print(f"📋 Transactions rows: {len(rows)}")
        
        count = 0
        for row in rows:
            if not any(row): continue
            d      = safe_str(row[0] if len(row) > 0 else "")
            if not d: continue
            ttype  = safe_str(row[1] if len(row) > 1 else "")
            desc   = safe_str(row[2] if len(row) > 2 else "")
            s_d     = safe_int(row[3] if len(row) > 3 else 0)
            g_d     = safe_int(row[4] if len(row) > 4 else 0)
            s_b     = safe_int(row[5] if len(row) > 5 else 0)
            g_b     = safe_int(row[6] if len(row) > 6 else 0)
            note   = safe_str(row[7] if len(row) > 7 else "")
            
            conn.execute("""
                INSERT INTO token_transactions (date, type, description, silver_delta, gold_delta, silver_balance, gold_balance, note)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (d, ttype, desc, s_d, g_d, s_b, g_b, note))
            count += 1
        
        conn.commit()
        print(f"✅ Imported {count} transactions")
        
        # Recalculate balances
        rebuild_balances(conn)
    else:
        print("⚠️  No 'Transactions' sheet found")
    
    # 5. RedeemItems
    if "RedeemItems" in sheet_map:
        ws = sheet_map["RedeemItems"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        print(f"📋 RedeemItems rows: {len(rows)}")
        
        count = 0
        for row in rows:
            if not any(row): continue
            conn.execute("""
                INSERT OR REPLACE INTO redeem_items (item_id, label, description, coin_type, cost, active, sort_order)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                safe_str(row[0] if len(row) > 0 else ""),
                safe_str(row[1] if len(row) > 1 else ""),
                safe_str(row[2] if len(row) > 2 else ""),
                safe_str(row[3] if len(row) > 3 else ""),
                safe_int(row[4] if len(row) > 4 else 0),
                1 if safe_bool(row[5] if len(row) > 5 else True) else 0,
                safe_int(row[6] if len(row) > 6 else 0)
            ))
            count += 1
        
        conn.commit()
        print(f"✅ Imported {count} redeem items")
    else:
        print("⚠️  No 'RedeemItems' sheet found")
    
    # 6. AppConfig
    if "AppConfig" in sheet_map:
        ws = sheet_map["AppConfig"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        print(f"📋 AppConfig rows: {len(rows)}")
        
        count = 0
        for row in rows:
            if not any(row): continue
            key = safe_str(row[0] if len(row) > 0 else "")
            val = safe_str(row[1] if len(row) > 1 else "")
            if key and val:
                conn.execute("INSERT OR REPLACE INTO app_config (key, value) VALUES (?, ?)", (key, val))
                if key == "exchange_rate":
                    conn.execute("UPDATE tokens SET exchange_rate=?", (safe_int(val),))
                count += 1
        
        conn.commit()
        print(f"✅ Imported {count} app configs")

def rebuild_balances(conn):
    """Rebuild silver_balance/gold_balance from transactions"""
    rows = conn.execute("""
        SELECT silver_delta, gold_delta FROM token_transactions
        ORDER BY id ASC
    """).fetchall()
    s_bal, g_bal = 0, 0
    for r in rows:
        s_bal += r["silver_delta"] or 0
        g_bal += r["gold_delta"] or 0
    
    conn.execute("UPDATE tokens SET silver_balance=?, gold_balance=?", (s_bal, g_bal))
    conn.commit()
    print(f"💰 Recalculated balances: silver={s_bal}, gold={g_bal}")

def verify(conn):
    print("\n=== 数据验证 ===")
    tables = ["evaluations", "activity_logs", "token_transactions", "redeem_items"]
    for table in tables:
        cnt = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
        print(f"  {table}: {cnt} 条")
    t = conn.execute("SELECT silver_balance, gold_balance, exchange_rate FROM tokens LIMIT 1").fetchone()
    if t:
        print(f"  tokens: 银币={t['silver_balance']}, 金币={t['gold_balance']}, 汇率={t['exchange_rate']}")
    cfg = conn.execute("SELECT value FROM app_config WHERE key='exchange_rate'").fetchone()
    if cfg:
        print(f"  app_config exchange_rate: {cfg['value']}")

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 deploy/import_xlsx_to_sqlite.py <xlsx_path> [db_path]")
        print("  Default db_path: infra/web/backend/data/poc.db")
        sys.exit(1)
    
    xlsx_path = os.path.abspath(sys.argv[1])
    if not os.path.exists(xlsx_path):
        print(f"❌ File not found: {xlsx_path}")
        sys.exit(1)
    
    # Default DB path: local dev path
    default_db = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                              "infra", "web", "backend", "data", "poc.db")
    db_path = sys.argv[2] if len(sys.argv) > 2 else default_db
    
    print(f"📂 XLSX: {xlsx_path}")
    print(f"🗄️  DB:   {db_path}")
    
    # Ensure db dir exists
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    
    # Load workbook
    print("\n📖 加载 Excel...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    print(f"✅ Sheets found: {wb.sheetnames}")
    
    # Connect to DB
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    
    # Clear and reimport
    ensure_tables(conn)
    
    print("\n--- 清空现有数据 ---")
    conn.executescript("""
        DELETE FROM evaluations;
        DELETE FROM activity_logs;
        DELETE FROM token_transactions;
        DELETE FROM redeem_items;
        DELETE FROM api_log;
        UPDATE tokens SET silver_balance=0, gold_balance=0, exchange_rate=5;
        DELETE FROM app_config;
        INSERT OR IGNORE INTO app_config (key, value) VALUES ('exchange_rate', '5');
    """)
    conn.commit()
    print("✅ 所有表已清空")
    
    print("\n=== 开始导入 ===")
    import_sheet(conn, wb)
    
    verify(conn)
    
    conn.close()
    wb.close()
    print("\n✅ 恢复完成！")

if __name__ == "__main__":
    main()